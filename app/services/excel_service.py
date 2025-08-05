# app/services/excel_service.py

import io
import logging
import pandas as pd
from typing import List, Dict, Optional
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

logger = logging.getLogger(__name__)

class ExcelService:
    @staticmethod
    async def parse_excel_file(file, required_columns: List[str]) -> List[Dict]:
        """
        Parse Excel file and return list of dictionaries.
        Supports both .xlsx and .xls formats.
        """
        try:
            # Read the uploaded file content
            content = await file.read()
            excel_buffer = io.BytesIO(content)
            
            # Use pandas to read Excel, specifying the sheet name for data
            df = pd.read_excel(excel_buffer, sheet_name="Stock Import Template", engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
            
            # Clean column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            # Validate required columns
            missing_columns = [col for col in required_columns if col.lower().replace(' ', '_') not in df.columns]
            if missing_columns:
                found_columns = ', '.join(df.columns)
                raise ValueError(f"Missing required columns in 'Stock Import Template' sheet: {', '.join(missing_columns)}. Found columns: {found_columns}. Make sure to upload a data file with the correct sheet and headers, not the instructions sheet.")
            
            # Convert to list of dicts, handling NaN values
            records = df.replace({pd.NA: None, float('nan'): None}).to_dict(orient='records')
            
            logger.info(f"Successfully parsed {len(records)} records from Excel file")
            return records
            
        except ValueError as ve:
            logger.error(f"Excel validation error: {str(ve)}")
            raise
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise ValueError(f"Invalid Excel file format or missing 'Stock Import Template' sheet: {str(e)}. Please use the downloaded template and fill in the data sheet.")

    @staticmethod
    def create_streaming_response(excel_data: io.BytesIO, filename: str) -> StreamingResponse:
        """Create streaming response for Excel download"""
        headers = {
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        return StreamingResponse(
            iter([excel_data.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

class StockExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Product Name",
        "Quantity",
        "Unit",
        "HSN Code",
        "Part Number",
        "Unit Price",
        "GST Rate",
        "Reorder Level",
        "Location"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for stock import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = StockExcelService.REQUIRED_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Steel Bolt M8x50",  # Product Name
            100,                 # Quantity
            "PCS",               # Unit
            "73181590",          # HSN Code
            "SB-M8-50",          # Part Number
            25.50,               # Unit Price
            18.0,                # GST Rate
            50,                  # Reorder Level
            "Warehouse A-1"      # Location
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Stock Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Product Name is mandatory and must be unique per organization.",
            "3. If product doesn't exist, it will be created automatically.",
            "4. Quantity must be a non-negative number.",
            "5. Unit Price and GST Rate should be numbers.",
            "6. Reorder Level should be an integer.",
            "7. Location is optional.",
            "8. Do not modify the header row.",
            "9. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Stock Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_stock(stock_data: List[Dict]) -> io.BytesIO:
        """Export stock data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Product Name", "Quantity", "Unit", "HSN Code", "Part Number",
            "Unit Price", "GST Rate", "Reorder Level", "Location"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(stock_data, 2):
            row_data = [
                item.get("product_name"),
                item.get("quantity"),
                item.get("unit"),
                item.get("hsn_code"),
                item.get("part_number"),
                item.get("unit_price"),
                item.get("gst_rate"),
                item.get("reorder_level"),
                item.get("location")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Stock data exported successfully: {len(stock_data)} records")
        return excel_buffer

class VendorExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Name",
        "Contact Number",
        "Email",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Pin Code",
        "State Code",
        "GST Number",
        "PAN Number"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for vendor import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendor Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = VendorExcelService.REQUIRED_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Test Vendor Inc",     # Name
            "1234567890",          # Contact Number
            "test@vendor.com",     # Email
            "123 Vendor Street",   # Address Line 1
            "Suite 456",           # Address Line 2
            "Vendor City",         # City
            "Vendor State",        # State
            "123456",              # Pin Code
            "27",                  # State Code
            "27AACFV1234D1Z5",     # GST Number
            "AACFV1234D"           # PAN Number
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Vendor Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Name and Contact Number are mandatory.",
            "3. Email, Address Line 2, GST Number, and PAN Number are optional.",
            "4. Pin Code and State Code must be valid.",
            "5. If vendor name exists, it will be updated; otherwise, created.",
            "6. Do not modify the header row.",
            "7. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Vendor Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_vendors(vendors_data: List[Dict]) -> io.BytesIO:
        """Export vendors data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendors Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Name", "Contact Number", "Email", "Address Line 1", "Address Line 2",
            "City", "State", "Pin Code", "State Code", "GST Number", "PAN Number"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(vendors_data, 2):
            row_data = [
                item.get("name"),
                item.get("contact_number"),
                item.get("email"),
                item.get("address1"),
                item.get("address2"),
                item.get("city"),
                item.get("state"),
                item.get("pin_code"),
                item.get("state_code"),
                item.get("gst_number"),
                item.get("pan_number")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Vendors data exported successfully: {len(vendors_data)} records")
        return excel_buffer

class CustomerExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Name",
        "Contact Number",
        "Email",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Pin Code",
        "State Code",
        "GST Number",
        "PAN Number"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for customer import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = CustomerExcelService.REQUIRED_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Test Customer LLC",   # Name
            "9876543210",          # Contact Number
            "test@customer.com",   # Email
            "456 Customer Ave",    # Address Line 1
            "Apt 789",             # Address Line 2
            "Customer City",       # City
            "Customer State",      # State
            "654321",              # Pin Code
            "29",                  # State Code
            "29AAACC1234E1Z7",     # GST Number
            "AAACC1234E"           # PAN Number
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Customer Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Name and Contact Number are mandatory.",
            "3. Email, Address Line 2, GST Number, and PAN Number are optional.",
            "4. Pin Code and State Code must be valid.",
            "5. If customer name exists, it will be updated; otherwise, created.",
            "6. Do not modify the header row.",
            "7. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Customer Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_customers(customers_data: List[Dict]) -> io.BytesIO:
        """Export customers data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Name", "Contact Number", "Email", "Address Line 1", "Address Line 2",
            "City", "State", "Pin Code", "State Code", "GST Number", "PAN Number"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(customers_data, 2):
            row_data = [
                item.get("name"),
                item.get("contact_number"),
                item.get("email"),
                item.get("address1"),
                item.get("address2"),
                item.get("city"),
                item.get("state"),
                item.get("pin_code"),
                item.get("state_code"),
                item.get("gst_number"),
                item.get("pan_number")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Customers data exported successfully: {len(customers_data)} records")
        return excel_buffer

class ProductExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Product Name",
        "HSN Code",
        "Part Number",
        "Unit",
        "Unit Price",
        "GST Rate",
        "Is GST Inclusive",
        "Reorder Level",
        "Description",
        "Is Manufactured"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for product import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers (including optional initial stock columns)
        headers = ProductExcelService.REQUIRED_COLUMNS + ["Initial Quantity", "Initial Location"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Test Product",        # Product Name
            "123456",              # HSN Code
            "TP-001",              # Part Number
            "PCS",                 # Unit
            100.0,                 # Unit Price
            18.0,                  # GST Rate
            "FALSE",               # Is GST Inclusive
            10,                    # Reorder Level
            "Test description",    # Description
            "FALSE",               # Is Manufactured
            50,                    # Initial Quantity (optional)
            "Warehouse A"          # Initial Location (optional)
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Product Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Product Name and Unit are mandatory.",
            "3. HSN Code, Part Number, Description are optional but recommended.",
            "4. Unit Price, GST Rate, Reorder Level should be numbers.",
            "5. Is GST Inclusive and Is Manufactured should be TRUE/FALSE or YES/NO.",
            "6. Initial Quantity and Initial Location are optional for initial stock setup.",
            "7. If product name exists, it will be updated; otherwise, created.",
            "8. Do not modify the header row.",
            "9. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Product Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_products(products_data: List[Dict]) -> io.BytesIO:
        """Export products data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Products Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Product Name", "HSN Code", "Part Number", "Unit",
            "Unit Price", "GST Rate", "Is GST Inclusive", "Reorder Level",
            "Description", "Is Manufactured"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(products_data, 2):
            row_data = [
                item.get("product_name"),
                item.get("hsn_code"),
                item.get("part_number"),
                item.get("unit"),
                item.get("unit_price"),
                item.get("gst_rate"),
                item.get("is_gst_inclusive"),
                item.get("reorder_level"),
                item.get("description"),
                item.get("is_manufactured")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Products data exported successfully: {len(products_data)} records")
        return excel_buffer