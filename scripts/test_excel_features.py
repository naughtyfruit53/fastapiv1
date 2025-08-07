#!/usr/bin/env python3
"""
Excel Feature Verification Script

This script verifies that all Excel import/export functionality is working correctly.
Run this script to test all Excel features without requiring a full database setup.

Usage: python scripts/test_excel_features.py
"""

import sys
import os
import io
import asyncio
from pathlib import Path

# Add the app directory to the Python path
sys.path.insert(0, str(Path(__file__).parent.parent))

def test_all_excel_services():
    """Test all Excel services can be imported and used"""
    try:
        from app.services.excel_service import (
            StockExcelService, ProductExcelService, 
            VendorExcelService, CustomerExcelService, ExcelService
        )
        print("✅ All Excel services imported successfully")
        return True
    except ImportError as e:
        print(f"❌ Failed to import Excel services: {e}")
        return False

def test_template_generation():
    """Test template generation for all modules"""
    try:
        from app.services.excel_service import (
            StockExcelService, ProductExcelService, 
            VendorExcelService, CustomerExcelService
        )
        
        print("\n📋 Testing template generation...")
        
        services = [
            ("Stock", StockExcelService),
            ("Product", ProductExcelService),
            ("Vendor", VendorExcelService),
            ("Customer", CustomerExcelService)
        ]
        
        all_passed = True
        for name, service in services:
            try:
                template = service.create_template()
                size = len(template.getvalue())
                print(f"✅ {name} template: {size} bytes")
            except Exception as e:
                print(f"❌ {name} template failed: {e}")
                all_passed = False
        
        return all_passed
    except Exception as e:
        print(f"❌ Template generation test failed: {e}")
        return False

def test_export_functionality():
    """Test export functionality for all modules"""
    try:
        from app.services.excel_service import (
            StockExcelService, ProductExcelService, 
            VendorExcelService, CustomerExcelService
        )
        
        print("\n📤 Testing export functionality...")
        
        # Sample data for each module
        test_data = {
            "Stock": [{
                "product_name": "Test Product",
                "hsn_code": "12345678",
                "part_number": "TP001",
                "unit": "PCS",
                "unit_price": 100.0,
                "gst_rate": 18.0,
                "reorder_level": 10,
                "quantity": 50,
                "location": "Warehouse A"
            }],
            "Product": [{
                "product_name": "Test Product",
                "hsn_code": "12345678",
                "part_number": "TP001",
                "unit": "PCS",
                "unit_price": 100.0,
                "gst_rate": 18.0,
                "is_gst_inclusive": False,
                "reorder_level": 10,
                "description": "Test description",
                "is_manufactured": False
            }],
            "Vendor": [{
                "name": "Test Vendor",
                "contact_number": "1234567890",
                "email": "test@vendor.com",
                "address1": "123 Test St",
                "address2": "",
                "city": "Test City",
                "state": "Test State",
                "pin_code": "123456",
                "state_code": "27",
                "gst_number": "27ABCDE1234F1Z5",
                "pan_number": "ABCDE1234F"
            }],
            "Customer": [{
                "name": "Test Customer",
                "contact_number": "9876543210",
                "email": "test@customer.com",
                "address1": "456 Test Ave",
                "address2": "",
                "city": "Test City",
                "state": "Test State",
                "pin_code": "654321",
                "state_code": "29",
                "gst_number": "29AAACC1234E1Z7",
                "pan_number": "AAACC1234E"
            }]
        }
        
        services = [
            ("Stock", StockExcelService.export_stock),
            ("Product", ProductExcelService.export_products),
            ("Vendor", VendorExcelService.export_vendors),
            ("Customer", CustomerExcelService.export_customers)
        ]
        
        all_passed = True
        for name, export_func in services:
            try:
                export_data = export_func(test_data[name])
                size = len(export_data.getvalue())
                print(f"✅ {name} export: {size} bytes")
            except Exception as e:
                print(f"❌ {name} export failed: {e}")
                all_passed = False
        
        return all_passed
    except Exception as e:
        print(f"❌ Export functionality test failed: {e}")
        return False

def test_import_parsing():
    """Test import/parsing functionality"""
    try:
        from app.services.excel_service import ExcelService, StockExcelService
        from openpyxl import Workbook
        
        print("\n📥 Testing import/parsing functionality...")
        
        # Create test Excel file
        test_data = [{
            "Product Name": "Test Product 1",
            "HSN Code": "12345678",
            "Part Number": "TP001",
            "Unit": "PCS",
            "Unit Price": 100.0,
            "GST Rate": 18.0,
            "Reorder Level": 10,
            "Quantity": 50,
            "Location": "Warehouse A"
        }]
        
        # Create Excel file with proper structure
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Import Template"
        
        # Add headers
        headers = list(test_data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, (key, value) in enumerate(row_data.items(), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Mock file object
        class MockFile:
            def __init__(self, buffer):
                self.buffer = buffer
                self.filename = "test.xlsx"
            
            async def read(self):
                return self.buffer.getvalue()
        
        # Test parsing
        async def test_parse():
            mock_file = MockFile(excel_buffer)
            records = await ExcelService.parse_excel_file(
                mock_file, 
                StockExcelService.REQUIRED_COLUMNS,
                "Stock Import Template"
            )
            return records
        
        records = asyncio.run(test_parse())
        if records and len(records) > 0:
            print(f"✅ Parsing successful: {len(records)} records")
            print(f"   Sample: {records[0]['product_name']}")
            return True
        else:
            print("❌ Parsing failed: No records returned")
            return False
            
    except Exception as e:
        print(f"❌ Import/parsing test failed: {e}")
        return False

def check_dependencies():
    """Check if required dependencies are installed"""
    print("🔍 Checking dependencies...")
    
    required_packages = ['pandas', 'openpyxl', 'fastapi']
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package)
            print(f"✅ {package}")
        except ImportError:
            print(f"❌ {package} - Missing")
            missing_packages.append(package)
    
    if missing_packages:
        print(f"\n❌ Missing dependencies: {', '.join(missing_packages)}")
        print("Install with: pip install " + " ".join(missing_packages))
        return False
    
    return True

def main():
    """Run all Excel functionality tests"""
    print("🧪 Excel Feature Verification")
    print("=" * 50)
    
    if not check_dependencies():
        return False
    
    tests = [
        ("Service Import", test_all_excel_services),
        ("Template Generation", test_template_generation),
        ("Export Functionality", test_export_functionality), 
        ("Import/Parsing", test_import_parsing)
    ]
    
    passed = 0
    total = len(tests)
    
    for test_name, test_func in tests:
        print(f"\n🧪 Running {test_name} test...")
        if test_func():
            passed += 1
        else:
            print(f"💥 {test_name} test failed")
    
    print("\n" + "=" * 50)
    print(f"📊 Results: {passed}/{total} tests passed")
    
    if passed == total:
        print("🎉 All Excel features are working correctly!")
        print("\n✨ You can now use Excel import/export functionality in your application.")
        return True
    else:
        print("⚠️  Some tests failed. Check the output above for details.")
        print("💡 Make sure all dependencies are installed and the app structure is correct.")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)